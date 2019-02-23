#!/usr/bin/env python
#
# Transform read counts from readcount.sh into Excel file.
#
# Assume script is fed all file names.
# Any file name ending in ".counttable" gets its own sheet 
# In count tables use '@' to distinguish single read from contig
#

import collections
import itertools
import operator
import os
import re
import sys
import warnings

import openpyxl

sys.path.append(os.path.join(sys.path[0], '../lib/python'))
from SURPIviz import SampleSheet

BARCODE = 0; COUNT = 1
SURPI_NOTE = "Note: all numbers are pre-calculated by SURPI"
NA = 'n/a'
# Sample sheet to be read must be of format:
# Barcode	Sample
# 1	JB_UX_02
# 2	2
# 3	JB_UX_03
# 4	4
# default location of sample sheet
SAMPLESHEET_V1 = '%s.samplesheet.txt'
SAMPLESHEET_V2 = '%s.SampleSheet.csv'

def logHeader():
	import os.path, sys, time
	return "%s\t%s\t" % (time.strftime("%a %b %d %H:%M:%S %Z %Y"), os.path.basename(sys.argv[0]))

def version():
	import os.path
	import sys
	print os.path.basename(sys.argv[0]), 'v1.1'
	sys.exit(0)

# ignore if cell value doesn't look like a parameter
reCellParam = re.compile(r'^\s*\[\[(\w+)]]\s*$')
def getCellParam(cell):
	if not isinstance(cell.value, basestring) or not cell.value.startswith('[['):
		return

	m = reCellParam.match(cell.value)
	if m is None:
		return

	return m.group(1)

# writes the barcode column headings
def barcode(sheet, cell):
	for i, barcode in enumerate(sheet.barcodes):
		c = cell.offset(row=0, column=i)
		c.value = addSampleToBarcode(sampleDict, barcode)
		copyCellStyle(cell, c)

	c = cell.offset(row=0, column=i+1)
	c.value = "Total"
	copyCellStyle(cell, c)

# TODO too much inconsistency around barcodes
reBarcodes = [
	re.compile(r'^[12]:\w+:\w+:(?P<barcode>(\+|\w)+)$'), 		# e.g., 1:N:0:63
	re.compile(r'^bar#@?(?P<barcode>(\+|\w)+)(/|(/[12]))?$'), 	# e.g., bar#GCCAAT or bar#GCCAAT/ or bar#GCCAAT/2
	re.compile(r'^#?(?P<barcode>(\+|\w)+)/?'), 					# e.g., #TGACCA or #TGACCA/
]
def extractBarcode(barcode):
	"""Return just the barcode from the various extant embeddings, if possible."""
	for reBarcode in reBarcodes:
		m = reBarcode.match(barcode)
		if m is not None:
			break
	else:
		return barcode

	return m.group('barcode')

def addSampleToBarcode(sampleDict, barcode):
	"""Return the barcode with sample prepended, if available."""
	if not sampleDict:
		return barcode

	extractedBarcode = extractBarcode(barcode)
	sample = sampleDict.get(extractedBarcode)
	if sample is None:
		return barcode

	return "%s %s" % (sample, extractedBarcode)

def copyCellStyle(cell, newCell):
	if cell.has_style:
		newCell.font = cell.font.copy()
		newCell.border = cell.border.copy()
		newCell.fill = cell.fill.copy()
		newCell.number_format = cell.number_format
		newCell.protection = cell.protection.copy()
		newCell.alignment = cell.alignment.copy()

# writes the formula for calculating humanmatched = preprocessed - humanunmatched
def humanmatched(sheet, cell):
	sum = 0
	for i in range(len(sheet.barcodes)):
		try:
			count = int(getHumanmatched(sheet, i))
		except KeyError:
			count = NA
		else:
			sum += count

		c = cell.offset(row=0, column=i)
		c.value = count
		copyCellStyle(cell, c)

	c = cell.offset(row=0, column=i+1)
	c.value = sum
	copyCellStyle(cell, c)

def getHumanmatched(sheet, i):
	pp = sheet.dataDict['preprocessed_barcode_count'][i][COUNT]
	hu = sheet.dataDict['humanunmatched_barcode_count'][i][COUNT]
	return int(pp) - int(hu)

# writes the formula humanmatched / preprocessed * 100
def pcthuman(sheet, cell):
	hmSum = ppSum = 0
	for i in range(len(sheet.barcodes)):
		c = cell.offset(row=0, column=i)
		try:
			hm = float(getHumanmatched(sheet, i))
			hmSum += hm
			pp = float(sheet.dataDict['preprocessed_barcode_count'][i][COUNT])
			ppSum += pp
		except KeyError:
			c.value = NA
		else:
			if pp:
				c.value = hm / pp * 100.0
			else:
				c.value = NA
		copyCellStyle(cell, c)

	c = cell.offset(row=0, column=i+1)
	if ppSum:
		c.value = hmSum / ppSum * 100.0
	else:
		c.value = NA
	copyCellStyle(cell, c)

# writes the formula preprocessed / rawdata * 100
def pctpreprocessed(sheet, cell):
	ppSum = rdSum = 0
	for i in range(len(sheet.barcodes)):
		c = cell.offset(row=0, column=i)
		try:
			pp = float(sheet.dataDict['preprocessed_barcode_count'][i][COUNT])
			ppSum += pp
			rd = float(sheet.dataDict['rawdata_barcode_count'][i][COUNT])
			rdSum += rd
		except KeyError:
			c.value = NA
		else:
			if rd:
				c.value = pp / rd * 100.0
			else:
				c.value = NA
		copyCellStyle(cell, c)

	c = cell.offset(row=0, column=i+1)
	if rdSum:
		c.value = ppSum / rdSum * 100.0
	else:
		c.value = NA
	copyCellStyle(cell, c)

class SummarySheet(object):
	"""Summary of various counts by barcode."""

	templateMap = {
		'barcode': barcode,
		'rawdata_barcode_count': r'%(base)s\.fastq$',
		'preprocessed_barcode_count': r'%(base)s\.preprocessed\.fastq$',
		'humanmatched_barcode_count': humanmatched,
		'humanunmatched_barcode_count': r'%(base)s\.(preprocessed\..+|.+\.bt2)\.unmatched\.fastq$',
		'ntmatched_barcode_count': '%(base)s\.NT\.snap\.matched\..+\.all\.annotated(\.sorted)?$',
		'ntvirusmatched_barcode_count': r'%(base)s\.NT\.snap\.matched\..*fl\.Viruses\.annotated$',
		'ntbacteriamatched_barcode_count': r'%(base)s\.NT\.snap\.matched\.(d\d\d\.)?fl.Bacteria.annotated$',
		'ntfungalmatched_barcode_count': r'%(base)s\.NT\.snap\.matched\.(d\d\d\.)?fl.Fungi.annotated$',
		'ntparasitematched_barcode_count': r'%(base)s\.NT\.snap\.matched\.(d\d\d\.)?fl.Parasite.annotated$',
		'ntplantsmatched_barcode_count': r'%(base)s\.NT\.snap\.matched\.(d\d\d\.)?fl.Plants.annotated$',
		'ntarthropodamatched_barcode_count': r'%(base)s\.NT\.snap\.matched\.(d\d\d\.)?fl.Arthropoda.annotated$',
		'ntnonchordateukmatched_barcode_count': r'%(base)s\.NT\.snap\.matched\.(d\d\d\.)?fl.nonChordatEuk.annotated$',
		'ntnonmammalchordatamatched_barcode_count': r'%(base)s\.NT\.snap\.matched\.(d\d\d\.)?fl.nonMammalChordat.annotated$',
		'ntnonprimatemammalmatched_barcode_count': r'%(base)s\.NT\.snap\.matched\.(d\d\d\.)?fl.nonPrimMammal.annotated$',
		'ntunmatched_barcode_count': r'%(base)s\.NT\.snap\.unmatched\.sam$',
		'nrrapsearchvirus_barcode_count': r'%(base)s\.Contigs\.(and\.NTunmatched\.Viral\.)?RAPSearch.*\.Viruses\.annotated$',
		'pct_human': pcthuman,
		'pct_preprocessed': pctpreprocessed,
	}

	def __init__(self, ws):
		self.ws = ws # openpyxl worksheet object
		self.dataDict = {}
		self.barcodes = None

	def populate(self, wb):
		print "%spopulating read count file '%s'" % (
				logHeader(), wb.readCountFile)
		fileNames, rowDict = self.readData(wb.readCountFile)
		self.normalizeBarcodes(fileNames, rowDict)
		self.parseData(rowDict)
		self.collate()

	def readData(self, filePath):
		# need original processing order of file names for normalizing
		# this way the first file has all the barcodes
		fileNames = []
		rowDict = collections.defaultdict(list)
		if inputDir:
			filePath = os.path.join(inputDir, filePath)
		with open(filePath, 'rU') as f:
			for line in f:
				try:
					fileName, barcode, count = line.split()
				except ValueError:
					print "%sERROR: bad line format: '%s'" % (logHeader(), line)
					raise
				fileNames.append(fileName)
				rowDict[fileName].append((barcode, count))
		return fileNames, rowDict

	def normalizeBarcodes(self, fileNames, rowDict):
		firstFile = ''
		for fileName in fileNames:
			# per Samia sort the columns
			rowDict[fileName].sort()
			# normalize the barcode representations
			barcodeCounts = rowDict[fileName]
			barcodeCounts = [(extractBarcode(barcode), count) for barcode, count in barcodeCounts]
			rowDict[fileName] = barcodeCounts
			barcodes = map(self.getBarcode, barcodeCounts)
			if not self.barcodes:
				self.barcodes = barcodes
				firstFile = fileName
			elif self.barcodes != barcodes:
				barcodeCounts = self.fillMissingBarcodes(fileName, barcodes, barcodeCounts)
				barcodes = map(self.getBarcode, barcodeCounts)
				if self.barcodes != barcodes:
					print "%sbarcodes reported for file %s differ from those reported for file %s: %s != %s" % (
							logHeader(), fileName, firstFile, barcodes, self.barcodes)

	# In existing datasets it is possible for barcodes not to survive various
	# processing steps, i.e., have a zero value, in which case they are absent
	# from the totals.
	def fillMissingBarcodes(self, fileName, barcodes, barcodeCounts):
		for barcode in self.barcodes:
			if barcode not in barcodes:
				print "%sbarcodes reported for file %s missing barcode '%s': assuming zero value" % (
						logHeader(), fileName, barcode)
				barcodeCounts.append((barcode, 0))
		barcodeCounts.sort()
		return barcodeCounts

	# collect barcodes
	getBarcode = operator.itemgetter(BARCODE)
	def parseData(self, rowDict):
		# build data dictionary of values only; ignore functions for now
		for cellParam, cellValue in self.templateMap.items():
			# simple lookup using regexp
			if isinstance(cellValue, basestring):
				formatDict = {'base': base}
				reHeading = cellValue % formatDict
				for heading in rowDict:
					# have seen differences in case over time, e.g. RAPsearch vs RAPSearch
					if re.match(reHeading, heading, flags=re.IGNORECASE) is not None:
						barcodeCounts = rowDict[heading]
						break
				else:
					print "%sno data found for template parameter '%s' using '%s': skipping" % (
							logHeader(), cellParam, reHeading)
					continue

				self.dataDict[cellParam] = barcodeCounts

	reCellParam = re.compile(r'^\s*\[\[(\w+)]]\s*$')
	def collate(self):
		for row in self.ws.rows:
			for cell in row:
				cellParam = getCellParam(cell)
				if cellParam is None:
					# not a cell parameter
					continue

				cellValue = self.templateMap.get(cellParam)
				if cellValue is None:
					print "%sno data source found for template parameter '%s' in sheet '%s'" % (
							logHeader(), cellParam, self.ws.title)
					continue

				# collate stats into rows for output
				if not isinstance(cellValue, basestring):
					# callable to calculate values
					cellValue(self, cell)
				else:
					# simple lookup, which we did in parse if it exists
					# possible that template parameter not in data file
					barcodeCounts = self.dataDict.get(cellParam)
					if barcodeCounts is None:
						print "%sno data found for parameter '%s' in worksheet '%s'" % (
								logHeader(), cellParam, self.ws.title)
						barcodeCounts = [(barcode, NA) for barcode in self.barcodes]

					# duplicate of parseData 
					# verify barcode order so we don't emit incorrect data
					barcodes = map(self.getBarcode, barcodeCounts)
					if self.barcodes != barcodes:
						print "%sfound inconsistency in barcodes: %s != %s" % (
								logHeader(), self.barcodes, barcodes)
						continue

					self.fillRow(cell, barcodeCounts)

		c = self.ws.cell(row=cell.row+2, column=1)
		c.value = SURPI_NOTE

	def fillRow(self, cell, barcodeCounts):
		sum = 0 # totals column
		for i, barcodeCount in enumerate(barcodeCounts):
			barcode, count = barcodeCount
			try:
				count = int(count)
			except ValueError:
				count = NA
			else:
				sum += count
			c = cell.offset(row=0, column=i)
			c.value = count
			copyCellStyle(cell, c)

		c = cell.offset(row=0, column=i+1)
		c.value = sum
		copyCellStyle(cell, c)


class CountTableSheet(object):
	"""Detail count table."""

	def __init__(self, ws):
		self.ws = ws # openpyxl worksheet object from template file
		self.templateCellDict = {}
		self.labels = None # species, genus, family
		self.barcodes = None
		self.headings = None
		self.dataRows = None

	def populate(self, wb):
		self.parseTemplate()
		for filePath in wb.countTableFiles:
			if inputDir:
				filePath = os.path.join(inputDir, filePath)
			fileName = os.path.basename(filePath)
			print "%spopulating count table file '%s'" % (
					logHeader(), filePath)
			try:
				headings, dataRows = self.readData(filePath)
			except IOError:
				print "%sWARNING: file not found: '%s'" % (logHeader(), filePath)
				continue
			self.parseData(fileName, headings, dataRows)
			self.collate(wb, fileName)

	def parseTemplate(self):
		"""Build a dictionary of parameterized cells."""
		for row in self.ws.rows:
			for cell in row:
				cellParam = getCellParam(cell)
				if cellParam is None:
					# not a cell parameter
					continue

				self.templateCellDict[cellParam] = cell

	# Samia wants columns sorted by barcode; we assume count tables already sorted.
	def readData(self, filePath):
		"""Read a count table and return headings and data rows."""
		headings = []
		dataRows = []
		with open(filePath, 'rU') as f:
			for line in f:
				if not headings:
					headings = [s.strip() for s in line.split('\t')]
					continue
				dataRows.append([s.strip() for s in line.split('\t')])
		return headings, dataRows

	def parseData(self, fileName, headings, dataRows):
		"""Validate and organize the count table data into taxa and counts."""
		# older heading might contain (@=contigbarcode)
		# newer contains tag label and no 'bar#' prefix
		# separate out the barcodes
		for i, heading in enumerate(headings):
			if heading.startswith('bar#'):
				self.labels = headings[:i]
				self.barcodes = headings[i:]
				break
		else:
			for i, heading in enumerate(headings):
				if heading == 'tag':
					self.labels = headings[:i+1]
					self.barcodes = headings[i+1:]
					break
			else:
				print "%s%s does not look like a count table" % (
						logHeader(), fileName)
				return

		# validate the data
		countStartIndex = len(self.labels)
		for data in dataRows:
			if len(data) != len(headings):
				print "%sline '%s' does not look like a count table" % (
						logHeader(), data)
				return

			# convert counts to integers
			for i in range(countStartIndex, len(data)):
				data[i] = float(data[i])

		self.headings = headings
		self.dataRows = dataRows

	def collate(self, wb, fileName):
		"""Apply count table data to new sheet using hard-coded positions and styles from template."""
		if not self.dataRows:
			return

		# row and column are 1-based
		# openpyxl.exceptions.SheetTitleException: Maximum 31 characters allowed in sheet title
		ws = wb.wb.create_sheet(title=self.getSheetName(fileName))
		cell = ws.cell(row=1, column=1)
		cell.value = fileName
		copyCellStyle(self.templateCellDict['filename'], cell)

		# headings consist of labels and barcodes
		# labels: species, genus, etc.
		labelCt = len(self.labels)
		for i, label in enumerate(self.labels):
			cell = ws.cell(row=2, column=i+1)
			cell.value = label
			copyCellStyle(self.templateCellDict['label_heading'], cell)

		# barcodes
		for i, barcode in enumerate(self.barcodes):
			cell = ws.cell(row=2, column=labelCt+i+1)
			cell.value = addSampleToBarcode(sampleDict, barcode)
			copyCellStyle(self.templateCellDict['barcode'], cell)

		# FIXME per Samia leave totals off for now
		# barcode total
		#cell = ws.cell(row=2, column=labelCt+i+2)
		#cell.value = "Total"
		#cell.style = self.templateCellDict['barcode'].style

		# rows of count data
		for i, data in enumerate(self.dataRows):
			sum = 0 # totals column
			for j, datum in enumerate(data):
				cell = ws.cell(row=i+3, column=j+1)
				cell.value = datum
				if j < labelCt:
					copyCellStyle(self.templateCellDict['label_value'], cell)
				else:
					copyCellStyle(self.templateCellDict['barcode_count'], cell)
					sum += datum

			# FIXME per Samia leave totals off for now
			# NOTE can have both contig (bar#@CGA) and barcode (bar#CGA)
			# How does this affect totals??
			#cell = ws.cell(row=i+3, column=j+2)
			#cell.value = sum
			#cell.style = self.templateCellDict['barcode_count'].style

		cell = ws.cell(row=i+5, column=1)
		cell.value = SURPI_NOTE

	def getSheetName(self, fileName):
		"""Return a shortened but useful name for worksheets."""
		return fileName.replace(base, '').replace('annotated', '').replace('counttable', '').replace('.', '')[:30]


class SummaryWorkbook(object):

	sheetDict = {
		'Summary': SummarySheet,
		'Count Table Template': CountTableSheet,
	}

	def __init__(self, readCountFile, countTableFiles):
		self.readCountFile = readCountFile
		self.countTableFiles = countTableFiles
		self.wb = None

	def readTemplate(self, fileName):
		self.wb = openpyxl.load_workbook(filename=fileName)

	def populate(self):
		names = self.wb.get_sheet_names()
		for name in names:
			# ws will be None if name does not exist but this should not happen
			ws = self.wb.get_sheet_by_name(name)
			if ws.title in self.sheetDict:
				print "%sprocessing template worksheet '%s'" % (
						logHeader(), ws.title)
				sheet = self.sheetDict[ws.title](ws)
				sheet.populate(self)

	def clean(self):
		"""Perform final housekeeping."""
		for name in self.sheetDict:
			# remove template sheets
			if name.endswith("Template"):
				ws = self.wb.get_sheet_by_name(name)
				self.wb.remove_sheet(ws)

	def removeSheet(self, name):
		ws = self.wb.get_sheet_by_name(name)
		self.wb.remove_sheet(ws)

	def write(self, fileName):
		self.wb.save(fileName)


def findSampleSheetFile(fileName):
	for filePath in [fileName, SAMPLESHEET_V2 % base, SAMPLESHEET_V1 % base]:
		if filePath is None:
			continue

		if inputDir:
			filePath = os.path.join(inputDir, filePath)
		if os.path.exists(filePath):
			break
		else:
			print "%ssample sheet %s not found" % (
					logHeader(), filePath)
	else:
		print "%sno sample sheet found: will report barcodes instead of sample names" % (
				logHeader(),)
		filePath = None

	return filePath


# Sample sheet might have been generated manually or programmatically.
# Use what we find but not essential.
# Format: Barcode	Sample
sampleDict = {}
def readSampleSheet(fileName):
	if fileName is None:
		return 

	print "%susing sample sheet '%s'" % (logHeader(), fileName)
	global sampleDict
	if fileName.endswith('.csv'):
		ss = SampleSheet.readV2(fileName)
		sampleDict = ss.data['sample_name']
	if not sampleDict and fileName.endswith('.txt'):
		sampleDict = SampleSheet.readV1(fileName)

	if not sampleDict:
		print "%ssample sheet '%s' has unexpected format: will report barcodes instead of sample names" % (logHeader(), fileName)


def usage(msg=None):
	print "Usage: %s [--version] [-d debug] [--input input directory] [--output output directory] [-s samplesheet file] [-r readcount file] <base identifier> <template file> [<counttable file>...]" % sys.argv[0]
	print "  input directory: source of count table files (default: current directory)"
	print "  output directory: destination of Excel file (default: current directory)"
	print "  samplesheet file: optional, one of two formats:"
	print "    Illumina samplesheet, comma-separated-values file (default: <base>.SampleSheet.csv)"
	print "    Old-style, two-column, tab-delimited file mapping sample names to barcodes"
	print "         column headings are 'Barcode\tSample' (default: <base>.samplesheet.txt)"
	print "  readcount file: BarcodeR1R2.log file containing summary of read counts"
	print "Translates barcode counts into Excel file using Excel template file."
	print "Will also translate one or more count tables of any type, e.g., GI, species, genus, family"
	print "Assumes count column headings begin with 'bar#'"
	if msg is not None:
		print msg


if __name__ == '__main__':
	import getopt
	inputDir = ''
	outputDir = ''
	sampleSheetFile = None
	readCountFile = None
	debug = False
	options, args = getopt.getopt(sys.argv[1:], "dr:s:", ['debug', 'input=', 'output=', 'version'])
	try:
		for option, value in options:
			if option == '--version':
				version()
			elif option == '-s':
				sampleSheetFile = value
			elif option == '--input':
				inputDir = value
			elif option == '--output':
				outputDir = value
			elif option == '-r':
				readCountFile = value
			elif option in ('-d', '--debug'):
				debug = True
	except getopt.GetoptError, msg:
		usage(msg)
		sys.exit(2)

	if len(args) < 2:
		usage("insufficient arguments supplied")
		sys.exit(2)

	if not debug:
		warnings.simplefilter("ignore")

	base = args[0]
	templateFile = args[1]
	countTableFiles = args[2:]

	if readCountFile is None and not countTableFiles:
		usage("insufficient arguments supplied")
		usage("must generate either the summary or at least one count table")
		sys.exit(2)

	sampleSheetFile = findSampleSheetFile(sampleSheetFile)
	readSampleSheet(sampleSheetFile)
	print "%sgenerating Excel summary file for %s" % (logHeader(), base)
	wb = SummaryWorkbook(readCountFile, countTableFiles)
	wb.readTemplate(templateFile)
	if readCountFile is None:
		wb.removeSheet('Summary')
	wb.populate()
	wb.clean()
	outFile = os.path.join(outputDir, '%s.summary.xlsx' % base)
	print "%swriting output file '%s'" % (logHeader(), outFile)
	wb.write(outFile)
